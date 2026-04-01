"""
Campaign Score Ingestion — ingest real-world campaign performance data.

Supports three ingestion methods:
  1. Excel upload (parsed server-side)
  2. Manual form entry
  3. API pull (analytics connectors — GA4, LinkedIn, Mailchimp)

Usage::

    from zeta_ima.memory.scores import ingest_scores_manual, get_campaign_score
"""

from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime, timezone
from typing import Any

from zeta_ima.config import settings

log = logging.getLogger(__name__)

SCORES_CONTAINER = "campaign_scores"


async def ingest_scores_manual(
    team_id: str,
    campaign_id: str,
    workflow_id: str = "",
    metrics: dict[str, Any] | None = None,
    composite_score: float = 0.0,
    ingested_by: str = "",
    notes: str = "",
) -> str:
    """Ingest scores via manual form entry."""
    from zeta_ima.infra.document_store import get_document_store

    entry_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)

    ds = get_document_store()
    await ds.upsert(SCORES_CONTAINER, {
        "id": entry_id,
        "team_id": team_id,
        "campaign_id": campaign_id,
        "workflow_id": workflow_id,
        "source": "manual",
        "metrics": metrics or {},
        "composite_score": composite_score,
        "ingested_by": ingested_by,
        "notes": notes,
        "created_at": now.isoformat(),
    })

    log.info("Ingested manual scores %s for campaign=%s", entry_id, campaign_id)
    return entry_id


async def ingest_scores_excel(
    team_id: str,
    file_bytes: bytes,
    ingested_by: str = "",
) -> list[str]:
    """
    Parse an Excel file and ingest campaign scores.

    Expected columns: campaign_id, metric_name, metric_value, composite_score, notes
    Returns list of created entry IDs.
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed — cannot parse Excel files")
        return []

    from zeta_ima.infra.document_store import get_document_store
    ds = get_document_store()

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if not ws:
        return []

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [str(c.value).lower().strip() for c in list(ws.iter_rows(min_row=1, max_row=1))[0] if c.value]

    entry_ids: list[str] = []
    now = datetime.now(timezone.utc)

    # Group metrics by campaign_id
    campaigns: dict[str, dict] = {}
    for row in rows:
        if not row or len(row) < 2:
            continue
        data = dict(zip(headers, row))
        cid = str(data.get("campaign_id", "")).strip()
        if not cid:
            continue

        if cid not in campaigns:
            campaigns[cid] = {
                "metrics": {},
                "composite_score": 0.0,
                "notes": "",
            }

        metric_name = str(data.get("metric_name", "")).strip()
        metric_value = data.get("metric_value", 0)
        if metric_name:
            campaigns[cid]["metrics"][metric_name] = metric_value

        cs = data.get("composite_score")
        if cs and float(cs) > 0:
            campaigns[cid]["composite_score"] = float(cs)

        note = data.get("notes", "")
        if note:
            campaigns[cid]["notes"] = str(note)

    for cid, info in campaigns.items():
        entry_id = str(uuid.uuid4())
        await ds.upsert(SCORES_CONTAINER, {
            "id": entry_id,
            "team_id": team_id,
            "campaign_id": cid,
            "workflow_id": "",
            "source": "excel",
            "metrics": info["metrics"],
            "composite_score": info["composite_score"],
            "ingested_by": ingested_by,
            "notes": info["notes"],
            "created_at": now.isoformat(),
        })
        entry_ids.append(entry_id)

    log.info("Ingested %d campaign scores from Excel for team=%s", len(entry_ids), team_id)
    return entry_ids


async def ingest_scores_api(
    team_id: str,
    source: str,
    campaign_id: str = "",
    credentials: dict[str, str] | None = None,
) -> list[str]:
    """
    Pull scores from an analytics API.

    Supported sources: "mailchimp", "ga4", "linkedin"
    Returns list of created entry IDs.
    """
    from zeta_ima.infra.document_store import get_document_store
    ds = get_document_store()

    entry_ids: list[str] = []
    now = datetime.now(timezone.utc)

    try:
        if source == "mailchimp":
            from zeta_ima.integrations.analytics_pull import pull_mailchimp
            data = await pull_mailchimp(credentials or {}, campaign_id=campaign_id)
        elif source == "ga4":
            from zeta_ima.integrations.analytics_pull import pull_ga4
            data = await pull_ga4(credentials or {}, campaign_id=campaign_id)
        elif source == "linkedin":
            from zeta_ima.integrations.analytics_pull import pull_linkedin
            data = await pull_linkedin(credentials or {}, campaign_id=campaign_id)
        else:
            log.warning("Unknown analytics source: %s", source)
            return []

        for item in data:
            entry_id = str(uuid.uuid4())
            await ds.upsert(SCORES_CONTAINER, {
                "id": entry_id,
                "team_id": team_id,
                "campaign_id": item.get("campaign_id", campaign_id),
                "workflow_id": item.get("workflow_id", ""),
                "source": source,
                "metrics": item.get("metrics", {}),
                "composite_score": item.get("composite_score", 0.0),
                "ingested_by": f"api:{source}",
                "notes": "",
                "created_at": now.isoformat(),
            })
            entry_ids.append(entry_id)

    except ImportError:
        log.warning("Analytics pull module not available for source=%s", source)
    except Exception as e:
        log.error("API score ingestion failed for source=%s: %s", source, e)

    return entry_ids


async def get_campaign_score(
    team_id: str,
    campaign_id: str = "",
    workflow_id: str = "",
) -> dict[str, Any] | None:
    """Get the latest score for a campaign or workflow."""
    from zeta_ima.infra.document_store import get_document_store
    ds = get_document_store()

    filters: dict[str, Any] = {"team_id": team_id}
    if campaign_id:
        filters["campaign_id"] = campaign_id
    elif workflow_id:
        filters["workflow_id"] = workflow_id
    else:
        return None

    results = await ds.query(SCORES_CONTAINER, filters=filters, limit=1)
    return results[0] if results else None


async def get_score_trend(
    team_id: str,
    campaign_id: str = "",
    limit: int = 20,
) -> list[dict[str, Any]]:
    """Get score history for trend analysis."""
    from zeta_ima.infra.document_store import get_document_store
    ds = get_document_store()

    filters: dict[str, Any] = {"team_id": team_id}
    if campaign_id:
        filters["campaign_id"] = campaign_id

    return await ds.query(SCORES_CONTAINER, filters=filters, limit=limit)
